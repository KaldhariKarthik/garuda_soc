#!/usr/bin/env python3
"""
gen_verification_plan.py -- builds the two GARUDA verification plans.

Two separate, independently submittable workbooks:

  tb/core/GARUDA_Core_Verification_Plan.xlsx   -- Block I, the RV32IM pipeline
      0 Summary          the five phases, status, blockers
      1 Unit             phase 1 -- per-module, rtl/core
      2 Core Sanity TB   phase 2
      3 Core ISA / ISS   phase 3
      4 SoC no Core      phase 4  (blocked -- SoC RTL does not exist yet)
      5 Full SoC         phase 5  (blocked)
      Coverage / Infrastructure / Open Questions

  tb/dsu/GARUDA_DSU_Verification_Plan.xlsx     -- Block II, the DSU
      0 Summary          what the block is and why it is the highest risk
      1 DSU Tests        the full test plan
      Coverage / Infrastructure / Open Questions

Sheet numbering mirrors the coach's document ("1Tops RISC-V SoC Verification
Strategy v1") so the two can be read side by side without translation.

Each workbook cross-references the other rather than duplicating rows.

Regenerate both with:  python3 tools/gen/gen_verification_plan.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HDR = ["#", "Test", "What could go wrong / what it checks",
       "How I test it", "Expected result", "Pri", "Done"]

TITLE_F = Font(bold=True, size=14)
SUB_F = Font(italic=True, size=9, color="555555")
HDR_F = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2F4F6F")
SEC_F = Font(bold=True)
SEC_FILL = PatternFill("solid", fgColor="DCE6F1")
BLOCK_FILL = PatternFill("solid", fgColor="F2DCDB")
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
WIDTHS = [5, 26, 50, 38, 44, 6, 8]


def sheet(wb, name, title, subtitle, rows, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = name
    ws["A1"] = title
    ws["A1"].font = TITLE_F
    ws["A2"] = subtitle
    ws["A2"].font = SUB_F
    for i, w in enumerate(WIDTHS):
        ws.column_dimensions[chr(65 + i)].width = w
    for c, h in enumerate(HDR, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font, cell.fill, cell.border = HDR_F, HDR_FILL, THIN
    r = 5
    n = 0
    for row in rows:
        if isinstance(row, str):                       # section banner
            cell = ws.cell(row=r, column=1, value=row)
            cell.font = SEC_F
            fill = BLOCK_FILL if row.startswith("!") else SEC_FILL
            if row.startswith("!"):
                cell.value = row[1:]
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill
                ws.cell(row=r, column=c).border = THIN
        elif row is None:
            pass
        else:
            n += 1
            row = list(row)
            # a leading "!" on the Test cell marks a red row (confirmed RTL
            # finding / hard blocker). Strip the marker, keep the highlight.
            flagged = isinstance(row[0], str) and row[0].startswith("!")
            if flagged:
                row[0] = row[0][1:]
            vals = [n] + row
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment, cell.border = WRAP, THIN
                if flagged:
                    cell.fill = BLOCK_FILL
        r += 1
    ws.freeze_panes = "A5"
    return ws


def note(ws, lines):
    r = ws.max_row + 2
    for ln in lines:
        c = ws.cell(row=r, column=1, value=ln)
        c.font = SUB_F if ln.startswith("  ") else Font(bold=True, size=9)
        r += 1


# ===========================================================================
# 0 SUMMARY
# ===========================================================================
CORE_SUMMARY = [
    "Phase 1 — Unit (per-module inside the core)",
    ["Scope", "26 modules in rtl/core + 14 in rtl/dsu",
     "Directed + constrained-random per module, task-based SV TB",
     "All P0 rows pass; functional coverage goals met per block",
     "P0", "In progress"],
    ["Status today", "6 of 26 core modules have any TB at all",
     "ex_smoke, id_ex_fwd, exmem_ifid, pipe_ctrl, csr_file, trap_ctrl",
     "20 modules unverified; 0 modules have covergroups", "P0", ""],
    "Phase 2 — Core Sanity TB",
    ["Scope", "garuda_core_top + I/D memory models + IRQ stimulus",
     "Hand-written instruction sequences via $readmemh",
     "Every instruction class executes; hazards and traps behave",
     "P0", "Not started"],
    ["Blocker", "needs AHB-Lite memory models (not AXI)",
     "Cadence AHB VIP, or hand-written behavioural slave",
     "Wait states + ERROR responses injectable", "P0", ""],
    "Phase 3 — Core ISA / ISS lockstep",
    ["Scope", "riscv-arch-test + riscv-dv streams vs Spike",
     "Compile with riscv-gnu-toolchain, run both, diff logs",
     "Signature match on all rv32ui/rv32um; 1k+ dv seeds clean",
     "P0", "Not started"],
    ["Blocker", "Spike does not know Custom-0 or CLIC",
     "Extend Spike via custom-extension hook -> DSU golden model",
     "DSU instrs retire in Spike with correct GPR/acc effects",
     "P0", "Needs decision"],
    "Phase 4 — SoC without core",
    ["Scope", "interconnect + peripheral subsystem, UVM agent",
     "UVM AHB master agent driving the bus in place of the core",
     "Register coverage 100%; interrupt paths end-to-end",
     "P1", "BLOCKED"],
    ["!Blocker", "rtl/ahb, rtl/apb, rtl/uart, rtl/gpio, rtl/i2c, rtl/spi_*, "
     "rtl/timers, rtl/plic, rtl/clic, rtl/dma are ALL empty directories",
     "n/a — no RTL exists to verify",
     "Cannot start until SoC RTL is written", "P1", "BLOCKED"],
    "Phase 5 — Full SoC with core",
    ["Scope", "boot + bare-metal regressions + long runs",
     "C tests via riscv-gnu-toolchain, boot from BROM",
     "Boot success; full regression clean; no deadlocks",
     "P1", "BLOCKED"],
    ["!Blocker", "depends on Phase 4 and on the SoC RTL existing",
     "n/a", "Cannot start", "P1", "BLOCKED"],
    "Cross-cutting",
    ["DSU verification", "Block II — the reason the chip exists",
     "Golden model + generated vectors + directed TB (see DSU sheet)",
     "All P0 DSU rows pass against DSUModel cycle-accurate compare",
     "P0", "In progress"],
    ["Formal (JasperGold)", "control blocks whose state space simulation "
     "only samples",
     "FPV on pipe_ctrl, trap_ctrl, forward_unit, hazard_forward_unit",
     "Priority/mutex properties proven, not just simulated",
     "P1", "Not started"],
    ["Coverage closure", "signoff is coverage, not 'it ran once'",
     "Xcelium covergroups + vManager merge across all regressions",
     "Functional goals met; code coverage 95% with waivers",
     "P0", "Not started"],
]

# ===========================================================================
# 1 UNIT
# ===========================================================================
UNIT = [
    "IF stage — garuda_pc_gen.v",
    ["reset vector", "PC must come out of reset at RESET_VECTOR or the core "
     "fetches garbage from cycle 0",
     "release reset, sample first I-port address",
     "haddr == RESET_VECTOR (parameterised, default 0x1000_0000)", "P0", ""],
    ["sequential increment", "PC must advance by 4 only when IF actually "
     "accepts — double-increment loses an instruction silently",
     "hold stall_i high/low across accepts", "PC+4 exactly once per accept",
     "P0", ""],
    ["redirect priority", "trap must beat EX redirect must beat ID redirect; "
     "wrong order fetches from the wrong target after a trap",
     "assert all three redirect sources in the same cycle",
     "trap target wins, then EX, then ID", "P0", ""],
    ["no advance under backpressure", "PC advancing while downstream is "
     "stalled drops instructions",
     "stall_i high for N cycles with valid fetch", "PC frozen for all N", "P0", ""],
    "IF stage — garuda_iport_ahb_master.v",
    ["AHB address/data phase", "protocol violation hangs the bus",
     "drive HREADY low for 1..N cycles mid-transfer",
     "HTRANS/HADDR held stable through the wait; no new transfer started",
     "P0", ""],
    ["drop_pending hazard", "redirect arriving while a fetch is outstanding "
     "must discard the returning data, not inject it after the flush",
     "assert redirect_i with fetch_outstanding high",
     "data_valid_o stays low for the discarded beat", "P0", ""],
    ["HRESP ERROR", "a bus error must become an instruction access fault "
     "(cause 1), not a silent wrong instruction",
     "slave returns ERROR two-cycle response",
     "instr_fault_o asserted with the right PC", "P0", ""],
    "IF stage — garuda_prefetch_buffer.v",
    ["fill/drain", "buffer must never return a stale instruction after a "
     "redirect",
     "fill buffer, redirect, check next delivered instruction",
     "buffer flushed; next instr is from the redirect target", "P0", ""],
    ["full/empty corners", "overflow or underflow corrupts instruction order",
     "back-to-back accepts with stalled consumer",
     "no instruction lost or duplicated", "P0", ""],
    "ID stage — decode_control.v",
    ["all RV32I opcodes", "a wrong control bit is a silent wrong answer, and "
     "the wild-goose-chase class of bug",
     "sweep every legal opcode x funct3 x funct7 combination",
     "control bundle matches a reference table", "P0", ""],
    ["M-extension detect", "funct7 = 0000001 must select MUL/DIV/REM",
     "sweep all 8 M-ops", "mul_en asserted, correct funct3 forwarded", "P0", ""],
    ["writeback select", "ALU vs LOAD vs PC+4 confusion writes the wrong "
     "value to the register file",
     "one instruction of each class", "mem_to_reg / reg_write correct", "P0", ""],
    ["illegal detection", "an undecodable word must raise illegal, not "
     "execute as something else",
     "sweep unused opcodes and reserved funct combinations",
     "illegal_instr_dec asserted, all enables low", "P0", ""],
    ["Custom-0 passthrough", "DSU opcode must set dsu_en and not be treated "
     "as illegal by the base decoder",
     "all 10 Custom-0 funct5 values", "dsu_en high, illegal low", "P0", ""],
    "ID stage — imm_gen.v",
    ["I/S/B/U/J formats", "wrong immediate = wrong branch target or wrong "
     "address; classic silent failure",
     "directed vectors per format incl. all-ones and sign boundaries",
     "matches a reference immediate extractor", "P0", ""],
    ["sign extension", "unsigned-extending a negative immediate breaks "
     "backwards branches",
     "negative immediates in every format", "correctly sign-extended to 32b",
     "P0", ""],
    "ID stage — regfile.v",
    ["x0 hardwired", "a write to x0 that sticks corrupts everything downstream",
     "write non-zero to x0, read back", "reads 0", "P0", ""],
    ["same-cycle write/read", "the WB->ID bypass policy must be defined and "
     "implemented consistently (write-first vs read-first)",
     "write and read the same register in one cycle",
     "documented policy holds — write-first per Sec.7", "P0", ""],
    ["dual read ports", "rs1/rs2 crosstalk", "read two different registers "
     "simultaneously, including the same register on both ports",
     "both ports return the correct value", "P1", ""],
    "ID stage — branch_predict.v",
    ["static direction", "backward-taken / forward-not-taken policy must "
     "match what EX compares against, or every branch mispredicts",
     "backward and forward branch offsets",
     "predict_taken matches the documented static policy", "P0", ""],
    ["target computation", "wrong predicted target wastes the redirect",
     "sweep branch offsets", "target = PC + imm", "P0", ""],
    "ID stage — hazard_forward_unit.v / load_use",
    ["load-use detect", "missing a load-use stall reads stale register data",
     "LOAD followed immediately by a dependent op on rs1, rs2, and both",
     "load_use_stall asserted for exactly one cycle", "P0", ""],
    ["no false stall", "over-stalling on x0 or on a non-load kills performance "
     "and can mask real bugs",
     "LOAD with rd=x0 then dependent op; ALU op then dependent op",
     "no stall", "P1", ""],
    "EX — alu.v",
    ["ADD/SUB wrap", "overflow must wrap two's complement, not saturate",
     "0x7FFFFFFF+1, 0x80000000-1", "wraps, no trap", "P0", ""],
    ["shift amount masking", "RV32 masks the shift amount to 5 bits; not "
     "masking gives 0 where the ISA says otherwise",
     "shift amounts 0, 1, 31, 32, 63", "amount masked to [4:0]", "P0", ""],
    ["SRL vs SRA", "logical vs arithmetic confusion breaks signed division "
     "idioms", "negative operand, both shift ops",
     "SRA keeps sign, SRL zero-fills", "P0", ""],
    ["SLT/SLTU sign", "signed/unsigned compare confusion around the sign bit",
     "pairs straddling 0x80000000",
     "matches signed and unsigned semantics respectively", "P0", ""],
    "EX — mul32.v",
    ["MUL/MULH/MULHSU/MULHU", "the three high-half variants differ only in "
     "signedness and are very easy to get wrong",
     "corner operand matrix incl. 0x80000000 x 0x80000000",
     "matches a Python reference for all four", "P0", ""],
    ["DIV/REM by zero", "the ISA mandates specific results, not a trap",
     "divide and remainder by zero",
     "DIV=-1 (all ones), REM=dividend, per RISC-V spec", "P0", ""],
    ["DIV overflow", "-2^31 / -1 is the one overflow case in the ISA",
     "0x80000000 / 0xFFFFFFFF", "DIV=0x80000000, REM=0", "P0", ""],
    ["latency / backpressure", "multi-cycle units are where handshake bugs "
     "hide", "issue back-to-back multiplies with EX stalled",
     "deterministic or bounded latency; no lost result", "P0", ""],
    "EX — branch_unit.v",
    ["compare matrix", "a wrong compare sends control flow to the wrong place "
     "and looks like a random crash",
     "BEQ/BNE/BLT/BGE/BLTU/BGEU with rs1<rs2, ==, >",
     "full 6x3 matrix matches reference", "P0", ""],
    ["sign-bit boundary", "signed vs unsigned compare divergence",
     "operands straddling 0x80000000", "signed and unsigned differ correctly",
     "P0", ""],
    "EX — csr_rw.v",
    ["RW/RS/RC semantics", "read-modify-write must read the OLD value and "
     "write the new one in the same instruction",
     "CSRRW/S/C with known prior CSR value",
     "rd gets old value, CSR gets new", "P0", ""],
    ["rs1=x0 / uimm=0", "RS/RC with a zero operand must NOT write (else you "
     "corrupt read-only-ish CSRs by reading them)",
     "CSRRS/CSRRC with rs1=x0", "no write occurs", "P0", ""],
    ["immediate variants", "CSRRWI/SI/CI take a 5-bit uimm, not a register",
     "all three immediate forms", "uimm zero-extended, not sign-extended",
     "P0", ""],
    "EX — forward_unit.v",
    ["EX/MEM over MEM/WB", "when both stages can forward the same register, "
     "the YOUNGER value must win or you read stale data",
     "same rd in EX/MEM and MEM/WB, dependent op in EX",
     "EX/MEM forwarded", "P0", ""],
    ["x0 never forwarded", "forwarding a write to x0 injects garbage",
     "instruction with rd=x0 followed by a dependent read",
     "no forward; regfile 0 used", "P0", ""],
    ["reg_write gating", "forwarding from an instruction that isn't writing "
     "is a pure data-corruption bug",
     "store/branch (reg_write=0) followed by dependent op", "no forward",
     "P0", ""],
    ["FORMAL", "the select priority is a small combinational lattice — prove "
     "it rather than sample it",
     "JasperGold FPV: assert one-hot select and correct priority",
     "properties proven", "P1", ""],
    "EX — ex_stage.v (result mux + DSU boundary)",
    ["result mux priority", "Sec.9.2 erratum: csr_rdata > dsu > mul > alu. A "
     "wrong order breaks CSRRW read-modify-write",
     "issue CSR, DSU, MUL and ALU ops incl. overlapping enables",
     "documented priority holds", "P0", ""],
    ["dsu_rd_valid gating", "ex_mem must capture the DSU result on "
     "dsu_rd_valid, NOT on dsu_en (valid self-gates during a busy hold)",
     "DSU op that stalls (dsu_busy) then completes",
     "no capture while busy; capture exactly once on valid", "P0", ""],
    ["rd_out source", "the DSU owns rd when it writes back; taking rd from "
     "the base decoder instead writes the wrong register",
     "DSU readback op with rd != base-decoded rd",
     "rd_out = dsu_rd_addr when dsu_rd_valid", "P0", ""],
    "MEM — load_store_unit.v",
    ["byte enables", "wrong BE for address[1:0] corrupts neighbouring bytes — "
     "a classic silicon escape",
     "SB/SH/SW at every offset 0..3", "HSIZE and write-data lane correct",
     "P0", ""],
    ["store data alignment", "store data must be replicated into the right "
     "lane for sub-word stores",
     "SB/SH at all offsets", "hwdata lane matches address", "P0", ""],
    ["misalignment detect", "a misaligned access must be killed BEFORE the "
     "bus address phase commits",
     "LW at offset 1/2/3, LH at odd offset",
     "misaligned flagged, no bus transfer started", "P0", ""],
    "MEM — load_formatter.v",
    ["sign vs zero extend", "LB/LH vs LBU/LHU confusion is silent and common",
     "negative byte/half at every offset",
     "LB/LH sign-extend, LBU/LHU zero-extend", "P0", ""],
    ["offset extraction", "the wrong byte lane returns a plausible-but-wrong "
     "value", "all size x offset combinations",
     "correct lane selected for each", "P0", ""],
    "MEM — d_port_ahb_master.v",
    ["wait states", "the pipeline must freeze, not drop, during HREADY low",
     "random 0..N wait states on every access",
     "mem_stall asserted; transfer completes intact", "P0", ""],
    ["HRESP ERROR", "must become load/store access fault (cause 5/7) with the "
     "right mtval", "two-cycle ERROR response on load and on store",
     "err_pulse -> mem_exception_valid, cause 5 or 7", "P0", ""],
    ["one outstanding", "issuing a second transfer before the first completes "
     "violates the D-port contract",
     "back-to-back load/store with wait states", "strictly one at a time",
     "P0", ""],
    "CONTROL — pipe_ctrl.v",
    ["hold table", "the Sec.11.4 hold composition decides which stages freeze; "
     "a wrong entry either loses or duplicates an instruction",
     "assert each hold source alone and in combination",
     "matches the documented hold table", "P0", ""],
    ["flush beats hold", "a squashed instruction must not be held — it is "
     "being discarded anyway",
     "flush and stall on the same register in the same cycle",
     "flush wins per Sec.11.4", "P0", ""],
    ["redirect priority", "trap > EX > ID; wrong order resumes at the wrong PC",
     "all three redirects asserted together", "trap target selected", "P0", ""],
    ["mem_stall gating", "redirects must not fire mid-D-port-wait",
     "assert a redirect while mem_stall is high",
     "redirect deferred until the wait clears", "P0", ""],
    ["WFI drain-precision", "WFI must hold PC, IF/ID AND ID/EX so the younger "
     "instruction does not execute during sleep (fixed 2026-07-20)",
     "assert wfi_hold, observe the hold pattern",
     "1_0_1_0_1_0_0_0_0 — EX/MEM free so older ops drain", "P0", "PASS"],
    ["FORMAL", "hold/flush composition is a small, high-consequence lattice",
     "JasperGold FPV: no-lost-instruction and no-duplicate properties",
     "proven, not sampled", "P1", ""],
    "CONTROL — csr_file.v",
    ["read/write every CSR", "an unimplemented or mis-decoded CSR silently "
     "reads 0", "sweep the whole implemented address list",
     "each reads back what was written; RO ones reject writes", "P0", ""],
    ["illegal CSR", "access to an unimplemented CSR, or a write to a "
     "read-only one, must raise illegal (cause 2)",
     "unimplemented address; write to misa/mip/mintstatus/dsu_ovf",
     "illegal_csr asserted", "P0", ""],
    ["mstatus push/pop", "trap entry must push MIE->MPIE and clear MIE; MRET "
     "must pop. Getting this wrong nests interrupts incorrectly",
     "trap then MRET with MIE initially set",
     "MPIE/MIE sequence per Sec.14.4", "P0", ""],
    ["minstret accuracy", "counts architectural retirement, not register "
     "writes — stores and branches retire without writing (fixed 2026-07-20)",
     "run a mixed stream incl. stores, branches and a faulting load",
     "minstret == instructions retired; faulting load NOT counted", "P0", ""],
    ["mcycle", "must free-run every cycle including stalls",
     "count over a stalled window", "increments unconditionally", "P1", ""],
    ["mip.MTIP / mip.MEIP", "mip is hardware state; a writable mip is a spec "
     "violation and lets firmware fake interrupts",
     "drive mtip and clic pending; attempt to write mip",
     "bit7/bit11 track hardware; write raises illegal", "P0", ""],
    ["dsu_ovf read-clear", "reading 0xBC0 must clear the sticky DSU overflow "
     "flag", "set overflow, read the CSR twice",
     "first read 1, flag cleared, second read 0", "P0", ""],
    "CONTROL — trap_ctrl.v  (highest-consequence block in the core)",
    ["EX exception priority", "Sec.14.1 order 1>2>3>4>6>11 — wrong order "
     "reports the wrong mcause for a genuinely faulting instruction",
     "assert several EX exception sources simultaneously",
     "highest-priority cause wins", "P0", "PASS"],
    ["MEM outranks EX", "the older instruction must win; otherwise a younger "
     "fault retires an older faulting access",
     "MEM access fault + EX illegal in the same cycle",
     "cause 5/7 taken, mepc from MEM", "P0", "PASS"],
    ["exception beats interrupt", "taking an interrupt instead of a pending "
     "exception loses the exception permanently",
     "EX illegal + CLIC take_cond together", "exception taken", "P0", "PASS"],
    ["mtval per cause", "Sec.14.2: 1->PC, 2->instr, 3->PC, 4/6->addr, "
     "5/7->addr, 11/int->0. Wrong mtval breaks the handler",
     "one case per cause", "mtval matches the table", "P0", "PASS"],
    ["MTI entry", "machine timer must be takeable, ranked below CLIC "
     "external, gated by mstatus.MIE (added 2026-07-20)",
     "mtip+mtie with MIE set / clear; MTI vs CLIC together",
     "cause 0x80000007 at mtvec base; CLIC wins when both", "P0", "PASS"],
    ["WFI wake MIE-independent", "with MIE clear the hart must still resume "
     "after the WFI, not sleep forever",
     "WFI, then assert a pending interrupt with MIE=0",
     "wfi_hold releases; no trap taken", "P0", "PASS"],
    ["MRET", "must pop mstatus and redirect to mepc without entering a trap",
     "MRET after a trap", "mret pulse, no trap_enter, redirect to mepc",
     "P0", "PASS"],
    ["nested / back-to-back traps", "a trap arriving during trap entry is the "
     "classic lockup-that-does-not-reproduce",
     "assert a second trap source during the entry bubbles",
     "no lost trap, no double entry", "P0", ""],
    ["FORMAL", "this block is small, control-only, and the most dangerous "
     "thing in the core — simulation only samples its state space",
     "JasperGold FPV: mutual exclusion of causes, priority ladder, "
     "no-trap-lost", "properties proven", "P0", ""],
    "CONTROL — clic_ctrl.v",
    ["level arbitration", "a lower-level interrupt pre-empting a higher one "
     "breaks the whole priority scheme",
     "sweep irq level vs mintthresh vs mintstatus.mil",
     "take_cond only when level > threshold and > current", "P0", ""],
    ["shv vectoring", "selectable hardware vectoring must read the right mtvt "
     "entry", "shv set and clear",
     "vector target from mtvt table vs mtvec base", "P0", ""],
    ["ack handshake", "a missing or duplicated ack loses or re-takes an "
     "interrupt", "take an interrupt, observe ack",
     "exactly one ack pulse with the matching id", "P0", ""],
    "Pipeline registers — if_id / id_ex / ex_mem / mem_wb_reg",
    ["hold retains", "a held stage that does not retain corrupts the "
     "in-flight instruction", "stall for N cycles",
     "all fields unchanged for N cycles", "P0", "PASS"],
    ["flush bubbles", "a flushed stage must drive all write-enables low, not "
     "just change the opcode", "assert flush",
     "reg_write/mem_write/valid all low; NOP encoding", "P0", "PASS"],
    ["flush beats stall", "Sec.11.4 priority", "both asserted together",
     "bubble inserted", "P0", "PASS"],
    ["retire tag", "valid must be low in a bubble or minstret over-counts "
     "(added 2026-07-20)", "flush and stall sequences",
     "valid low in every bubble", "P0", ""],
    "Integration — garuda_core_top.v",
    ["elaboration with real DSU", "the frozen Sec.9.2 boundary must hold "
     "against real hardware, not just the stub",
     "make elab_core_dsu", "clean elaborate, zero warnings", "P0", "PASS"],
    ["EX equivalence stub vs DSU", "with dsu_en low the real DSU must be "
     "completely inert and inject no X into the result mux",
     "make test_ex and test_ex_dsu", "bit-identical results", "P0", "PASS"],
]

# ===========================================================================
# 2 CORE SANITY TB
# ===========================================================================
SANITY = [
    "Bring-up",
    ["reset & first fetch", "if the first fetch is wrong nothing else matters",
     "release reset, watch the I-port",
     "fetch from RESET_VECTOR, PC advances by 4", "P0", ""],
    ["I-MEM / D-MEM models", "the TB harness itself must be trustworthy "
     "before it can judge the DUT",
     "AHB-Lite behavioural slaves with programmable wait states",
     "zero-wait and N-wait both work", "P0", ""],
    "One of each instruction class",
    ["arithmetic", "the base sanity check", "ADDI/ADD/SUB/AND/OR/XOR sequence "
     "with known operands", "self-checking: final register values match", "P0", ""],
    ["branches", "control flow must actually redirect",
     "taken and not-taken of each branch type",
     "correct path executed", "P0", ""],
    ["jumps", "JAL/JALR link register and target",
     "JAL forward/backward, JALR with register base",
     "correct target; ra holds PC+4", "P0", ""],
    ["loads/stores", "round-trip through the D-port",
     "store then load back at all sizes and offsets",
     "value round-trips with correct extension", "P0", ""],
    ["M-extension", "each of the 8 M ops executes at least once",
     "MUL/MULH/MULHSU/MULHU/DIV/DIVU/REM/REMU",
     "results match reference", "P0", ""],
    ["CSR access", "privilege plumbing works end to end",
     "CSRRW/S/C on mscratch and mstatus", "read-modify-write correct", "P0", ""],
    ["DSU Custom-0", "the accelerator executes inside the real pipeline",
     "a short MAC sequence then readback",
     "register receives the expected accumulator value", "P0", ""],
    "Hazards in a real pipeline",
    ["RAW back-to-back", "forwarding under real timing",
     "dependent ops at distance 1, 2, 3", "correct values, no stall needed "
     "beyond load-use", "P0", ""],
    ["load-use", "the one hazard that needs a stall",
     "LOAD then immediate dependent use", "one bubble, correct data", "P0", ""],
    ["branch + hazard", "a branch whose condition depends on the previous op",
     "ADD then BEQ on the result", "correct direction taken", "P0", ""],
    ["DSU busy hold", "same-accumulator interlock in the real pipeline",
     "back-to-back DSU readbacks on one accumulator",
     "pipeline holds, no lost instruction", "P0", ""],
    "Traps and interrupts",
    ["ECALL / EBREAK", "synchronous trap entry from a real instruction stream",
     "execute each", "mepc/mcause/mtval correct, handler entered", "P0", ""],
    ["illegal instruction", "a bad word mid-stream",
     "inject 0xDEADBEEF", "cause 2, mtval = the instruction word", "P0", ""],
    ["misaligned load/store", "cause 4/6 from a real address",
     "LW at an odd address", "trap taken before the bus transfer", "P0", ""],
    ["bus error", "cause 5/7 from the D-port",
     "D-MEM model returns ERROR", "trap with the faulting address in mtval",
     "P0", ""],
    ["IRQ mid-stream", "an interrupt arriving at an arbitrary point must take "
     "at a clean boundary — this is where lockups come from",
     "assert CLIC irq at many different offsets into a loop",
     "trap at an instruction boundary; mepc resumable", "P0", ""],
    ["MRET resume", "the handler must return to exactly the right place",
     "trap then MRET", "execution resumes at mepc, state restored", "P0", ""],
    ["WFI sleep/wake", "drain-precise hold plus wake",
     "WFI then interrupt", "younger instruction did NOT execute during sleep",
     "P0", ""],
    "Memory timing stress",
    ["random wait states", "the pipeline must be indifferent to bus latency",
     "randomise I and D wait states 0..8 per access",
     "same architectural result as zero-wait", "P0", ""],
    ["flush + outstanding fetch", "redirect while a fetch is in flight",
     "branch mispredict with I-port wait states",
     "no stale instruction injected", "P0", ""],
]

# ===========================================================================
# 3 CORE ISA / ISS
# ===========================================================================
ISS = [
    "Toolchain bring-up",
    ["riscv-gnu-toolchain", "everything downstream needs it",
     "install riscv32-elf, compile and link a bare-metal hello",
     "test.elf -> test.hex loads and runs on the DUT", "P0", ""],
    ["Spike ISS", "the golden reference for the whole core",
     "build riscv-isa-sim configured for RV32IM, machine mode only",
     "Spike executes the same hex and logs commits", "P0", ""],
    ["log comparator", "the actual pass/fail engine",
     "script diffing DUT retire log vs Spike commit log per instruction",
     "first divergence reported with PC and instruction", "P0", ""],
    ["retire trace from RTL", "lockstep needs a per-instruction retire "
     "signal — the mem_wb retire tag added 2026-07-20 is exactly this",
     "dump PC/instr/rd/rd_value on retire_o",
     "one log line per retired instruction", "P0", ""],
    "Compliance",
    ["RISCOF setup", "the official conformance framework",
     "configure RISCOF with the DUT plugin and Spike as reference",
     "framework runs end to end on one test", "P0", ""],
    ["rv32ui", "base integer conformance",
     "run the full rv32ui arch-test suite",
     "all signatures match the reference", "P0", ""],
    ["rv32um", "M-extension conformance", "run the full rv32um suite",
     "all signatures match", "P0", ""],
    ["privileged / trap tests", "machine-mode trap behaviour",
     "the machine-mode subset of arch-test",
     "signatures match", "P0", ""],
    "Random instruction streams",
    ["riscv-dv bring-up", "the tool that finds what you did not think of",
     "configure riscv-dv for RV32IM machine-mode, Spike as reference",
     "one random program generated, compiled and compared", "P0", ""],
    ["hazard stress", "dense back-to-back dependencies",
     "riscv-dv with high dependency density",
     "1k seeds clean", "P0", ""],
    ["control stress", "dense branches and jumps",
     "riscv-dv branch-heavy configuration", "1k seeds clean", "P0", ""],
    ["exception stress", "random illegal and misaligned accesses",
     "riscv-dv with exception generation enabled",
     "1k seeds clean; all traps match Spike", "P0", ""],
    ["interrupt stress", "random IRQ injection during random code",
     "riscv-dv interrupt mode + CLIC stimulus",
     "no lockups; mepc always resumable", "P0", ""],
    "!GARUDA-specific ISS problems (need a decision — see Open Questions)",
    ["!Custom-0 in Spike", "Spike will trap DSU instructions as illegal, so "
     "any program containing one diverges immediately",
     "extend Spike via its custom-extension hook, calling the DSU golden "
     "model; or exclude DSU instrs from the compare",
     "DSU instructions retire in Spike with correct effects", "P0", ""],
    ["!CLIC in Spike", "base Spike models CLINT, not CLIC. GARUDA's entire "
     "trap architecture is CLIC-based",
     "decide: model CLIC in Spike, or verify CLIC traps by directed + "
     "formal only and exclude them from lockstep",
     "documented decision, agreed with the coach", "P0", ""],
    ["!mip layout", "GARUDA uses MEIP as a CLIC pending summary — compliance "
     "tests that poke mip/mie may flag it",
     "run the privileged suite and triage any mip/mie failures",
     "either conformant, or a documented deliberate deviation", "P1", ""],
    "Benchmarks",
    ["Dhrystone / CoreMark", "a long real program is a different stress than "
     "random streams", "compile and run to completion",
     "correct result; no deadlock; performance recorded", "P1", ""],
]

# ===========================================================================
# 4 / 5 SoC  (blocked)
# ===========================================================================
SOC_NO_CORE = [
    "!This entire phase is BLOCKED",
    ["!RTL does not exist", "rtl/ahb, rtl/apb, rtl/ahb2apb, rtl/uart, "
     "rtl/gpio, rtl/i2c, rtl/spi_master, rtl/spi_slave, rtl/pwm, rtl/timers, "
     "rtl/plic, rtl/clic, rtl/dma, rtl/brom, rtl/isram, rtl/dsram are all "
     "empty directories",
     "n/a", "cannot start until the SoC RTL is written", "P1", "BLOCKED"],
    "Planned scope (for when it unblocks)",
    ["bus agent", "the coach's template assumes AXI; GARUDA is AHB-Lite",
     "UVM AHB-Lite master agent (Cadence VIP) in place of the core",
     "agent drives the interconnect", "P1", ""],
    ["address map decode", "wrong decode sends a transfer to the wrong slave",
     "walk every slave's address range and the gaps between",
     "each access lands on the intended slave; gaps error cleanly", "P1", ""],
    ["register access", "RW/RO/W1C behaviour per peripheral register",
     "read/write every addressable register",
     "100% register coverage; field behaviour correct", "P1", ""],
    ["AHB-to-APB bridge", "protocol conversion and the clock-domain crossing "
     "(200MHz core vs 100MHz peripheral) is a classic escape",
     "randomised traffic across the bridge",
     "no lost or duplicated transfers; CDC clean", "P1", ""],
    ["interrupt routing", "a peripheral interrupt that never reaches the "
     "controller is invisible until silicon",
     "generate each peripheral event",
     "every source verified end-to-end to the controller input", "P1", ""],
    ["multi-master arbitration", "core I-port, core D-port and DMA all "
     "contend", "concurrent traffic from all masters",
     "no starvation, no protocol violation", "P1", ""],
]

FULL_SOC = [
    "!This entire phase is BLOCKED — depends on Phase 4",
    ["!prerequisite", "needs SoC RTL and a passing Phase 4",
     "n/a", "cannot start", "P1", "BLOCKED"],
    "Planned scope",
    ["boot from BROM", "if it does not boot, nothing else runs",
     "reset into the boot ROM", "reaches main()", "P1", ""],
    ["memory tests", "TCM integrity", "march and walking-ones over 384KB TCM",
     "no bit errors", "P1", ""],
    ["peripheral loopback", "end-to-end through real firmware",
     "UART/SPI/GPIO loopback tests in C", "data round-trips", "P1", ""],
    ["timer interrupt in C", "the MTI path exercised by real firmware",
     "set mtimecmp, enable, wait", "handler entered at the right time", "P1", ""],
    ["DSU firmware kernel", "the APF collision-avoidance kernel is the point "
     "of the chip", "run the real APF routine on real data",
     "output vector matches the Python reference", "P0", ""],
    ["long run stability", "deadlocks appear at hour scale, not test scale",
     "multi-million-cycle random + firmware soak",
     "no deadlock, no protocol stall beyond bounds", "P1", ""],
]

# ===========================================================================
# COVERAGE
# ===========================================================================
CORE_COVERAGE = [
    "The coach's signoff bar is coverage, not 'the test passed'",
    ["current state", "there are ZERO covergroups anywhere in the project "
     "today — every existing TB is a pass/fail $display",
     "audit tb/ for covergroup declarations",
     "this row closes when every P0 block has a covergroup", "P0", ""],
    "Functional coverage — per block",
    ["decode_control", "the biggest cross in the design",
     "cross opcode x funct3 x funct7", "all legal combinations hit", "P0", ""],
    ["load/store", "size x offset is where alignment bugs live",
     "cross size x address[1:0] x signed/unsigned",
     "all combinations hit", "P0", ""],
    ["ALU", "corner values, not just random",
     "bins on 0, 1, -1, 0x7FFFFFFF, 0x80000000; shift amounts 0/1/31/32+",
     "all bins hit", "P0", ""],
    ["hazards", "dependency patterns",
     "cross {rs1 hazard, rs2 hazard, both} x {ALU->use, LOAD->use, "
     "STORE->use} x distance 1/2/3", "all bins hit", "P0", ""],
    ["traps", "every cause, and the priority pairs",
     "bins per mcause; cross of simultaneous exception sources",
     "every implemented cause taken at least once", "P0", ""],
    ["CLIC", "level vs threshold vs current level",
     "cross irq level x mintthresh x mintstatus.mil x shv",
     "arbitration space covered", "P0", ""],
    ["AHB protocol", "wait-state and response buckets",
     "bins on wait-state count 0/1/2/>2 and HRESP OKAY/ERROR",
     "all buckets hit on both ports", "P0", ""],
    "Code coverage",
    ["statement/branch/toggle", "the coach's bar is 95% with waivers",
     "Xcelium code coverage across the full regression, merged in vManager",
     ">=95%, every exclusion justified in writing", "P0", ""],
    ["waiver review", "an unjustified waiver is a hole with paperwork",
     "review every exclusion with the coach",
     "each waiver has a documented reason", "P0", ""],
    "Regression",
    ["seed count", "one passing seed proves nothing",
     "riscv-dv N-seed regression, N per the coach's guidance (1k-10k)",
     "stable pass rate across N seeds", "P0", ""],
    ["bug burn-down", "signoff is closure, not test count",
     "track P0/P1 bugs found vs closed over time",
     "all P0/P1 closed; burn-down curve flat before tapeout", "P0", ""],
]

# ===========================================================================
# INFRASTRUCTURE
# ===========================================================================
CORE_INFRA = [
    "Simulators",
    ["Cadence Xcelium", "the signoff simulator",
     "xrun -f <filelist>; make SIM=xrun <target>",
     "available — all Cadence tools licensed", "P0", "HAVE"],
    ["Vivado xsim", "fallback used for the 2026-07-20 fixes because xrun was "
     "not on PATH on this machine",
     "make SIM=xsim <target>",
     "works; xrun path still needs one confirming run", "P1", "PARTIAL"],
    ["JasperGold", "formal for the control blocks",
     "FPV on pipe_ctrl, trap_ctrl, forward_unit",
     "available — high value, not yet used", "P1", "HAVE"],
    ["vManager", "coverage merge and regression management",
     "merge coverage across all phases",
     "available, not yet used", "P0", "HAVE"],
    "RISC-V toolchain (all open source, none installed yet)",
    ["riscv-gnu-toolchain", "compiles every test program",
     "riscv32-elf precompiled release; -elf for bare metal",
     "riscv32-unknown-elf-gcc on PATH", "P0", "TODO"],
    ["Spike (riscv-isa-sim)", "the golden ISS for lockstep",
     "build for RV32IM machine-mode; NEEDS Custom-0 + CLIC work",
     "runs the same hex as the DUT and logs commits", "P0", "TODO"],
    ["RISCOF", "compliance framework",
     "pip install riscof; DUT plugin + Spike reference plugin",
     "runs riscv-arch-test end to end", "P0", "TODO"],
    ["riscv-dv", "constrained-random instruction generator",
     "chipsalliance/riscv-dv, SV/UVM, Spike as reference",
     "generates and compares random streams", "P0", "TODO"],
    "Verification IP",
    ["AHB-Lite VIP", "do not hand-write bus models — the VIP's protocol "
     "assertions catch what a hand model allows",
     "Cadence AHB VIP on I-port and D-port",
     "protocol checking active in every core-level run", "P0", "TODO"],
    ["!unit framework from coach", "the strategy doc offers "
     "deeptech360/1tops_vyom_framework and an instruction encoder/decoder — "
     "building our own first means rewriting later",
     "ASK THE COACH for access before writing more unit TBs",
     "framework in hand and adopted", "P0", "ASK"],
    "GARUDA-owned models",
    ["DSU golden model", "the core cannot be lockstepped against Spike until "
     "Spike understands Custom-0, which needs this model",
     "tools/golden/DSU_Golden.py — see the separate DSU plan",
     "DONE — cycle-accurate, validated against RTL 2026-07-20", "P0", "HAVE"],
    ["APF reference", "the collision-avoidance kernel needs a reference too",
     "Python model of the APF routine",
     "used in Phase 5 firmware check", "P1", "TODO"],
]

# ===========================================================================
# OPEN QUESTIONS
# ===========================================================================
CORE_QUESTIONS = [
    "Decisions needed from the coach",
    ["Spike + Custom-0", "Spike traps DSU instructions as illegal, so lockstep "
     "diverges on any program containing one",
     "proposal: extend Spike via its custom-extension hook to call "
     "DSU_Golden.py's DSUModel",
     "agreed approach before Phase 3 starts", "P0", "ASK"],
    ["Spike + CLIC", "base Spike models CLINT, not CLIC; GARUDA's whole trap "
     "architecture is CLIC-based",
     "proposal: verify CLIC traps by directed + formal, exclude from lockstep",
     "agreed approach", "P0", "ASK"],
    ["CLIC vs PLIC", "both rtl/clic and rtl/plic exist as empty directories. "
     "The core is CLIC-based; the coach's reference SoC uses CLINT+PLIC",
     "ask which the SoC will actually carry",
     "documented decision — it changes the trap architecture", "P0", "ASK"],
    ["AXI vs AHB in the strategy doc", "phases 4 and 5 of the strategy assume "
     "AXI/AXI-Lite/APB; GARUDA is AHB-Lite. Section 2.8 also references OBI "
     "(gnt/rvalid) which GARUDA does not use",
     "confirm the AHB translation of those phases is acceptable",
     "agreed", "P1", "ASK"],
    ["unit framework", "the doc offers a framework; adopting it late means "
     "rewriting the unit TBs",
     "request deeptech360/1tops_vyom_framework access",
     "framework in hand before more unit TBs are written", "P0", "ASK"],
    ["formal is not in the strategy doc", "JasperGold is licensed and the "
     "control blocks are ideal FPV targets — worth adding as a layer",
     "propose FPV on pipe_ctrl, trap_ctrl, forward_unit",
     "agreed scope", "P1", "ASK"],
    "Schedule reality",
    ["!RTL freeze date passed", "README targets RTL freeze end of June; today "
     "is 2026-07-20. Core and DSU RTL are complete and elaborate together, so "
     "the RTL side is essentially met — but the slip is now entirely "
     "verification debt against a GDSII date of end-October",
     "re-plan with the coach using this workbook",
     "agreed, dated plan with owners", "P0", "ASK"],
    "Cross-reference",
    ["DSU has its own plan", "block II is verified separately; its open RTL "
     "findings (FLAG-A, FLAG-C) are tracked there",
     "see tb/dsu/GARUDA_DSU_Verification_Plan.xlsx",
     "both plans reviewed together", "P0", ""],
]


# ===========================================================================
# DSU  (block II) -- the original 35 rows, corrected, plus the gaps the
# 2026-07-20 golden-model audit exposed.
# ===========================================================================
DSU = [
    "MAC_SEL — multiply-accumulate (the core operation)",
    ["basic accumulate", "acc must equal running sum of a·b; a wrong sign or "
     "dropped term corrupts the force vector",
     "feed known a,b pairs incl 0, +1, -1, max, min; also random",
     "acc matches DSUModel", "P0", ""],
    ["2-cycle pipe latency", "REWRITTEN 2026-07-20: the product does not land "
     "'one cycle late', it lands on the next write_en to that accumulator — "
     "which may never arrive. A trailing op is required to commit it",
     "compute, then read the same acc on the next cycle, and again after a "
     "settle op", "matches DSUModel cycle for cycle, not just eventually",
     "P0", ""],
    ["three accs independent", "writing FX must not disturb FY or MAG",
     "interleave computes across FX/FY/MAG",
     "each acc holds only its own stream", "P0", ""],
    "MACSUB — subtract instead of add",
    ["subtract correctness", "two's-complement: both products inverted, needs "
     "+2 correction not +1",
     "subtract known values, both product signs",
     "acc decreases by exactly a·b", "P0", ""],
    ["+2 correction boundary", "if the correction is wrong the result is off "
     "by 1 or 2 every time", "directed cases at sign boundaries",
     "result exact, no off-by-one", "P0", ""],
    ["+2 relies on high lane = 0", "the +2 works only because operand_router "
     "forces the high product to zero for MACSUB. If MACSUB ever gained a dot "
     "form the correction would be wrong",
     "assert the high product is zero for every MACSUB",
     "pb == 0 always; documented as a design invariant", "P1", ""],
    "MACABS — absolute value then multiply",
    ["abs(0x8000) clamp", "−32768 has no positive twin in 16-bit; naive abs "
     "wraps back negative", "drive rs1 low half = 0x8000",
     "abs clamps to 0x7FFF (+32767), stays positive", "P0", ""],
    ["normal abs", "positive stays, negative flips", "sweep low-half ±",
     "|x| correct for all", "P1", ""],
    ["high lane forced 0", "MACABS is single-lane; the high product must not "
     "leak in", "random high halves", "high 16x16 contributes nothing",
     "P1", ""],
    "MACDOT — dot product (both lanes at once)",
    ["dual-lane sum", "acc += a0·b0 + a1·b1 in one issue",
     "packed 2×16 operands, random", "acc matches both-lane golden", "P0", ""],
    "MACLOAD — preset the accumulator",
    ["sign-extend 32→48", "loading a 32-bit value into a 48-bit acc must "
     "sign-extend, not zero-extend",
     "load 0x8000_0000, 0x7FFF_FFFF, random", "acc = sign-extended rs1",
     "P0", ""],
    "MACCLEAR — zero the accumulator",
    ["clears only selected acc", "must zero the chosen acc and leave the "
     "other two", "preload all three, clear one",
     "target = 0, others intact", "P0", ""],
    "MACSAT — clamp accumulator to 32-bit range",
    ["in-range passes", "if it fits in signed-32 it should pass through",
     "top 17 bits all-0 or all-1", "value unchanged", "P0", ""],
    ["positive clamp", "too big → clamp to +max, don't wrap",
     "acc just over +2^31-1", "acc = 0x0000_7FFF_FFFF", "P0", ""],
    ["negative clamp", "too small → clamp to −min", "acc just under −2^31",
     "acc = 0xFFFF_8000_0000", "P0", ""],
    ["exact boundary", "the boundary values themselves must NOT clamp",
     "acc = exactly ±2^31 edge", "pass through, overflow flag stays 0",
     "P0", ""],
    ["overflow flag set", "clamping means data was lost — the flag must "
     "record it", "out-of-range input", "sticky overflow = 1", "P0", ""],
    "MACRD_LO / MACRD_HI — read accumulator back to a register",
    ["RD_LO low 32 bits", "the register only takes 32 bits; the low half must "
     "come back exactly", "preload known acc, read low", "rd = acc[31:0]",
     "P0", ""],
    ["RD_HI sign-extend", "upper 16 bits sign-extended into a 32-bit register",
     "acc with top bit 0 and 1", "rd = sign-extended acc[47:32]", "P0", ""],
    ["read blocked during stall", "rd_valid must not fire while the unit is "
     "busy", "read right after a 2-cycle op on the same acc",
     "valid waits until busy clears", "P0", ""],
    "MACSHIFT — shift accumulator (scale the result)",
    ["arithmetic right (dir=0)", "signed right shift must keep the sign, not "
     "zero-fill", "negative acc, shift 1..47",
     "sign preserved, low 32 returned", "P0", ""],
    ["logical left (dir=1)", "left shift zero-fills", "shift 1..47",
     "zero-fill left, low 32 returned", "P0", ""],
    ["shift amount 48–63", "the field is 6 bits but the acc is only 48 wide",
     "amt = 48, 63",
     "RESOLVED: RTL and DSUModel agree (left→0, right→all sign bits). Needs "
     "an assembler contract, not an RTL fix — FLAG-D", "P1", ""],
    "Illegal instructions — must be rejected, not executed",
    ["acc_sel = 11", "only 3 accumulators exist; select 11 is garbage",
     "any op with acc_sel = 11", "illegal flagged, nothing changes", "P0", ""],
    ["unknown funct5", "opcodes 10–31 aren't real DSU ops",
     "sweep unused funct5", "illegal flagged", "P0", ""],
    ["bad funct3", "only 000 and 001 are valid formats",
     "funct3 = 010..111", "illegal flagged", "P0", ""],
    ["illegal changes nothing", "a rejected instruction must leave all state "
     "untouched", "illegal ops between good ones",
     "accs + flag unchanged", "P0", ""],
    ["I-type ignores funct5", "MACSHIFT is legal for ANY funct5 because the "
     "field is part of the immediate — confirm that is deliberate",
     "MACSHIFT with funct5 sweeping 0..31",
     "all accepted; documented as intended", "P1", ""],
    "Overflow flag — the check-engine light",
    ["sticky behaviour", "a 1-cycle overflow mid-calc must stay latched",
     "cause one overflow, run benign ops", "flag stays high", "P0", ""],
    ["clear by firmware", "a CSR read clears it", "pulse csr_clear_overflow",
     "flag → 0", "P0", ""],
    ["clear vs set same cycle", "if clear and overflow hit together, clear "
     "wins", "assert both in the same cycle",
     "flag = 0 — confirm this policy is intended (FLAG-E)", "P1", ""],
    ["overflow only on accumulate", "load/clear/saturate-writeback must not "
     "raise the accumulate overflow", "MACLOAD/MACCLEAR with extreme values",
     "mac overflow stays low (is_accum gating)", "P1", ""],
    "Reset & flush",
    ["async reset", "reset must zero all three accs, the flag, the pending "
     "product registers and the stall state",
     "assert rst_n mid-stream", "everything back to 0", "P0", ""],
    ["!flush semantics — ROW CORRECTED", "the old row expected 'acc = 0 after "
     "flush'. That is WRONG: the RTL zeroes only sum_reg/carry_reg and keeps "
     "acc, which is almost certainly right — a trap must not destroy "
     "accumulator state",
     "3*4, 5*6, flush, settle — confirmed in RTL 2026-07-20",
     "acc keeps its committed value (12); the pending product (30) is "
     "discarded. FLAG-B", "P0", "FIXED"],
    "!Pipeline structure — found by the 2026-07-20 golden-model audit",
    ["!FLAG-A: pending-product poisoning", "MACLOAD/MACCLEAR/MACSAT assert "
     "write_en, so they also latch sum_reg/carry_reg with products of their "
     "OWN operand fields. The next accumulate adds that stale product in. "
     "Masked today only because the generator drives rs1=rs2=0 for those ops",
     "MACCLEAR FY with rs1=7, rs2=9, then MAC_SEL 0*0, then settle",
     "CONFIRMED BUG: FY = 63 instead of 0. Needs an RTL fix or a documented "
     "ABI constraint", "P0", "OPEN"],
    ["!FLAG-C: MAC→readback not interlocked", "dsu_stall covers "
     "readback-after-readback, but readback-after-MULTIPLY is the hazard that "
     "actually loses data — the product is still pending in sum_reg",
     "MAC_SEL then immediate MACRD_LO on the same accumulator",
     "reads a stale accumulator. Extend dsu_stall, or require a drain op",
     "P0", "OPEN"],
    ["!FLAG-C: MACSAT sees stale acc", "MACSAT saturates cluster_out = acc, "
     "which excludes any pending product; the pending product is then added "
     "AFTER the clamp, so the result can exceed the saturation range",
     "MAC_SEL then immediate MACSAT on the same accumulator, then settle",
     "value outside the clamped range — confirm and fix", "P0", "OPEN"],
    ["stranded final product", "the last product of any sequence never "
     "commits until another op touches that accumulator",
     "run a MAC sequence and read without a settle op",
     "documented: either an RTL fix or a mandatory trailing drain", "P0", ""],
    "Integration boundary — dsu_top inside ex_stage",
    ["dsu_en = 0 fully inert", "the core relies on the DSU being silent for "
     "every non-Custom-0 instruction; today's stimulus format has no dsu_en "
     "field, so this is UNTESTED",
     "sweep every DSU instruction with dsu_en low",
     "no acc change, no rd_valid, no illegal, no overflow", "P0", ""],
    ["no X into the EX mux", "an X from the DSU propagates into ex_result and "
     "poisons the pipeline", "make test_ex_dsu (EX smoke vs real DSU)",
     "bit-identical to the stub run", "P0", "PASS"],
    ["dsu_busy holds the pipe", "the busy signal must stall EX without losing "
     "the instruction", "back-to-back same-acc readbacks in the real core",
     "pipeline holds, instruction retires once", "P0", ""],
    ["rd_addr ownership", "the DSU owns rd when it writes back",
     "DSU readback with a distinctive rd",
     "the right architectural register is written", "P0", ""],
    "Small building blocks (compare directly against the Python model)",
    ["operand_router", "abs16 incl 0x8000, lane gating",
     "exhaustive over 16-bit input", "matches DSU_Golden.abs16 / routing",
     "P0", ""],
    ["saturation_unit", "the in-range window + clamp constants",
     "boundary sweep", "matches the model", "P0", ""],
    ["barrel_shifter", ">>> vs << over all amounts", "full amount sweep 0..63",
     "matches the model", "P0", ""],
    ["csa_3to2 / kogge_stone_49", "sum/carry compression + the 49-bit final "
     "add; the overflow flag is derived from exactly this truncation",
     "random vectors vs the bit-literal model (NOT vs plain a+b+c)",
     "matches csa_3to2() and kogge_stone_49() in DSU_Golden.py", "P1", ""],
    "Methodology",
    ["cycle-accurate compare", "the old flow compared architectural values "
     "with a settle-op workaround, which hides exactly the class of bug "
     "FLAG-A turned out to be",
     "TB drives dsu_top and DSUModel in lockstep, comparing every cycle",
     "no divergence on any cycle of any test", "P0", ""],
    ["migrate the generator", "tools/gen/DSU_gen.py still uses the timeless "
     "ArchDSU and inserts settle ops",
     "port it to DSUModel and drop the drain hack",
     "vectors carry per-cycle expectations", "P0", ""],
    ["the missing testbench", "the golden model, generator and plan all "
     "exist; the SV testbench that consumes them does not",
     "write tb/dsu/tb_dsu_top.v reading the .mem files",
     "runs under Xcelium; this is the DSU critical path", "P0", ""],
]


# ===========================================================================
# DSU-specific summary / coverage / infrastructure / questions
# ===========================================================================
DSU_SUMMARY = [
    "What this block is",
    ["DSU = Block II", "3-MAC cluster, 16x16 signed multiply, 48-bit "
     "accumulators, 2-cycle CSA pipeline, 10 Custom-0 instructions. It exists "
     "for real-time APF collision avoidance — a wrong result here is a drone "
     "that does not avoid the obstacle",
     "n/a", "n/a", "P0", ""],
    ["why it is the highest risk in the chip", "it is the ONLY block with no "
     "external reference model. Spike knows RV32IM; it does not know "
     "Custom-0. If our model is wrong, nothing else catches it",
     "n/a", "n/a", "P0", ""],
    "Status today",
    ["RTL", "14 modules in rtl/dsu, complete and integrated into ex_stage",
     "make elab_core_dsu", "elaborates clean with the core", "P0", "PASS"],
    ["golden model", "reference for every expected value",
     "tools/golden/DSU_Golden.py — DSUModel (cycle-accurate) + ArchDSU",
     "DONE, validated against RTL 2026-07-20", "P0", "PASS"],
    ["vector generator", "directed + random stimulus",
     "tools/gen/DSU_gen.py -> dsu_stim.mem / dsu_expected.mem",
     "works; still uses ArchDSU + settle hack, needs migrating", "P0",
     "PARTIAL"],
    ["!testbench", "the missing link — nothing consumes the vectors",
     "tb/dsu/tb_dsu_top.v does not exist",
     "NOT WRITTEN. This is the DSU critical path", "P0", "TODO"],
    ["!RTL bugs open", "2 confirmed findings from the golden-model audit",
     "FLAG-A and FLAG-C, see the Open Questions sheet",
     "need a decision before RTL freeze", "P0", "OPEN"],
    "How this block is verified",
    ["layer 1 — sub-blocks", "operand_router, saturation_unit, "
     "barrel_shifter, csa/kogge compared directly against the Python model",
     "exhaustive or wide random sweeps",
     "bit-identical to the model", "P0", ""],
    ["layer 2 — instruction level", "each of the 10 Custom-0 instructions "
     "against DSUModel", "generated directed + random vectors",
     "no divergence on any cycle", "P0", ""],
    ["layer 3 — pipeline/hazard", "the 2-cycle structure, dsu_busy, flush, "
     "reset — where the real bugs turned out to be",
     "directed sequences targeting pending-product behaviour",
     "matches DSUModel cycle for cycle", "P0", ""],
    ["layer 4 — integration", "dsu_top inside ex_stage inside the core",
     "make test_ex_dsu; then core-level DSU sequences",
     "inert when dsu_en low; correct writeback when high", "P0", "PARTIAL"],
]

DSU_COVERAGE = [
    "Functional coverage — DSU",
    ["instruction x accumulator", "every op must be exercised on every "
     "accumulator, not just FX",
     "cross funct5 (0..9) x acc_sel (FX/FY/MAG)",
     "all 30 legal combinations hit", "P0", ""],
    ["operand corners", "sign boundaries are where 16-bit maths breaks",
     "bins on 0x0000, 0x0001, 0x7FFF, 0x8000, 0x8001, 0xFFFF for each lane",
     "all bins hit on both lanes", "P0", ""],
    ["accumulator range", "overflow and saturation edges",
     "bins on in-range, just-over +2^31, just-under -2^31, near 48-bit "
     "overflow", "all bins hit", "P0", ""],
    ["shift space", "amount x direction",
     "cross shift_amt bins (0, 1, 31, 32, 47, 48, 63) x direction",
     "all bins hit", "P0", ""],
    ["illegal space", "every rejection path taken",
     "bins on acc_sel=11, unknown funct5, bad funct3",
     "each illegal cause hit at least once", "P0", ""],
    ["pipeline sequences", "the hazard space the audit exposed",
     "cross previous-op x current-op x same/different accumulator",
     "MAC->read, MAC->sat, clear->MAC, load->MAC all hit", "P0", ""],
    ["flush / reset", "asserted at every pipeline position",
     "bins on flush during compute, during sat, while idle",
     "all positions hit", "P0", ""],
    "Code coverage",
    ["statement/branch/toggle", "the coach's bar is 95% with waivers",
     "Xcelium code coverage over the DSU regression, merged in vManager",
     ">=95% across rtl/dsu, every exclusion justified", "P0", ""],
    "Signoff",
    ["all P0 rows pass", "the bar for tapeout",
     "full DSU regression under Xcelium",
     "zero divergence vs DSUModel; all P0 closed", "P0", ""],
    ["FLAG closure", "no open RTL question at freeze",
     "review FLAG-A and FLAG-C with the coach",
     "each either fixed in RTL or documented as an ABI constraint", "P0", ""],
]

DSU_INFRA = [
    "Models we own (nobody else has these)",
    ["DSU_Golden.py — DSUModel", "cycle-accurate reference: models the 2-stage "
     "MAC pipeline, pending-product registers, dsu_en, flush, reset and the "
     "dsu_stall interlock. CSA and 49-bit adder are bit-literal, so overflow "
     "detection matches the RTL exactly",
     "tools/golden/DSU_Golden.py", "DONE — validated against RTL 2026-07-20",
     "P0", "HAVE"],
    ["DSU_Golden.py — ArchDSU", "timeless value model. Maths sanity only; it "
     "cannot judge timing and must not be used for RTL comparison",
     "same file", "kept for reference", "P1", "HAVE"],
    ["DSU_gen.py", "builds directed + random vectors as $readmemh files",
     "tools/gen/DSU_gen.py -> dsu_stim.mem (4 words/test), "
     "dsu_expected.mem (8 words/test)",
     "works; must migrate off ArchDSU and drop the settle hack", "P0",
     "PARTIAL"],
    ["!tb_dsu_top.v", "the testbench that consumes the vectors and drives the "
     "RTL — the one missing piece",
     "SV TB: $readmemh both files, drive dsu_top, compare every cycle",
     "NOT WRITTEN", "P0", "TODO"],
    "Tools",
    ["Cadence Xcelium", "signoff simulator", "xrun -f rtl/dsu/filelist.f",
     "available", "P0", "HAVE"],
    ["Vivado xsim", "fallback used on machines without Cadence tools",
     "xvlog/xelab", "works — used for the 2026-07-20 flag confirmation",
     "P1", "HAVE"],
    ["JasperGold", "the saturation and overflow logic is small, pure "
     "combinational maths — ideal FPV targets",
     "FPV on saturation_unit, operand_router, overflow_flag",
     "available, not yet used", "P1", "HAVE"],
    ["vManager", "coverage merge", "merge DSU regression coverage",
     "available, not yet used", "P0", "HAVE"],
    "Downstream dependency",
    ["Spike custom extension", "the CORE cannot be lockstepped against Spike "
     "until Spike understands Custom-0 — and the only thing that can teach it "
     "is this golden model",
     "extend Spike via its custom-extension hook, calling DSUModel",
     "DSU instructions retire correctly in Spike", "P0", "TODO"],
]

DSU_QUESTIONS = [
    "!Confirmed RTL findings — golden-model audit, 2026-07-20",
    ["!FLAG-A: pending-product poisoning", "in mac_unit.v, sum_reg/carry_reg "
     "and acc are written on the same edge, both gated by write_en, and "
     "acc_next uses the PREVIOUS sum_reg. So MACLOAD/MACCLEAR/MACSAT — which "
     "also assert write_en — latch the products of their OWN operand fields. "
     "The next accumulate adds that stale product in",
     "MACCLEAR FY with rs1=7, rs2=9, then MAC_SEL FY 0*0, then settle. "
     "Confirmed on real RTL under xsim",
     "GOT FY = 63, EXPECTED 0 (63 = 7x9). Fix: gate the sum_reg/carry_reg "
     "write on a real accumulate — or document that operand fields must be "
     "zero for these ops", "P0", "OPEN"],
    ["!FLAG-C: MAC->readback not interlocked", "dsu_stall interlocks "
     "readback-after-readback (RD_LO/RD_HI/MACSAT), but readback-after-"
     "MULTIPLY is the hazard that actually loses data, because the product is "
     "still sitting in sum_reg",
     "MAC_SEL then immediate MACRD_LO on the same accumulator",
     "reads a stale accumulator. Fix: extend dsu_stall to cover MAC->read, or "
     "mandate a drain instruction in the ABI", "P0", "OPEN"],
    ["!FLAG-C: MACSAT clamps a stale value", "MACSAT saturates cluster_out = "
     "acc, which excludes any pending product. The pending product is then "
     "added AFTER the clamp, so the final value can sit outside the "
     "saturation range the instruction was supposed to enforce",
     "MAC_SEL then immediate MACSAT on the same accumulator, then settle",
     "value escapes the clamp range — confirm and fix", "P0", "OPEN"],
    ["!stranded final product", "the last product of any sequence never "
     "commits until another op touches that accumulator. The generator hides "
     "this today by appending a settle op",
     "run a MAC sequence and read back without a settle",
     "either an RTL fix or a mandatory trailing drain in the ABI", "P0",
     "OPEN"],
    "!Plan corrections",
    ["!row 31 of the old plan was wrong", "it expected 'flush -> acc = 0'. The "
     "RTL zeroes only sum_reg/carry_reg and KEEPS acc — which is almost "
     "certainly correct, since a trap must not destroy accumulator state",
     "3*4, 5*6, flush, settle — confirmed on real RTL",
     "acc = 12 (kept), pending 30 discarded. Row corrected; confirm the "
     "intent was deliberate — FLAG-B", "P0", "FIXED"],
    "Behaviours to confirm as intended",
    ["FLAG-D: shift amount 48-63", "the field is 6 bits, the accumulator is "
     "48 wide. RTL and DSUModel agree (left -> 0, arithmetic right -> all "
     "sign bits), so this is not an RTL bug",
     "sweep amt 48..63 both directions",
     "needs an assembler/ABI contract stating the range is reserved", "P1",
     "ASK"],
    ["FLAG-E: clear vs overflow same cycle", "overflow_flag gives priority to "
     "csr_clear over a simultaneous overflow, so an overflow occurring in the "
     "same cycle as a firmware clear is lost",
     "assert both in one cycle",
     "confirm this is the intended policy", "P1", "ASK"],
    ["I-type ignores funct5", "MACSHIFT is legal for any funct5 value because "
     "those bits are part of the immediate",
     "MACSHIFT with funct5 sweeping 0..31",
     "confirm deliberate; it removes a whole class of illegal-instruction "
     "checking for I-type", "P1", "ASK"],
    ["dsu_en = 0 is untested", "the stimulus record format has no dsu_en "
     "field, so 'the DSU is inert for every non-DSU instruction' — which the "
     "whole core depends on — has never been checked",
     "add dsu_en to the stimulus format and sweep it low",
     "no state change, no rd_valid, no illegal, no overflow", "P0", "OPEN"],
    "Cross-reference",
    ["core has its own plan", "the DSU is verified as a block here; its "
     "integration into the pipeline is tracked in the core plan",
     "see tb/core/GARUDA_Core_Verification_Plan.xlsx",
     "both plans reviewed together", "P0", ""],
]


CORE_OUT = os.path.join(os.path.dirname(__file__), "..", "..", "tb", "core",
                        "GARUDA_Core_Verification_Plan.xlsx")
DSU_OUT = os.path.join(os.path.dirname(__file__), "..", "..", "tb", "dsu",
                       "GARUDA_DSU_Verification_Plan.xlsx")


def build_core():
    wb = openpyxl.Workbook()
    sheet(wb, "0 Summary", "GARUDA Core — Verification Plan",
          "Block I, the RV32IM pipeline. Phases mirror '1Tops RISC-V SoC "
          "Verification Strategy v1'. The DSU has its own plan: "
          "tb/dsu/GARUDA_DSU_Verification_Plan.xlsx",
          CORE_SUMMARY, first=True)

    ws = sheet(wb, "1 Unit", "Phase 1 — Unit Level (module inside the core)",
               "Strategy doc §2. One section per module in rtl/core. "
               "'Done' = PASS means a testbench exists and passes today.",
               UNIT)
    note(ws, ["Legend:",
              "  Pri  P0 = must pass to tape out.  P1 = should pass.",
              "  Done PASS = TB exists and passes.  blank = not written yet.",
              "Status: 6 of 26 core modules have any testbench today.",
              "The strategy doc has no unit section for CSR / trap / "
              "privilege — for GARUDA those are the highest-consequence "
              "blocks, so they are added here."])

    ws = sheet(wb, "2 Core Sanity TB", "Phase 2 — Core Sanity Testbench",
               "Strategy doc §3. garuda_core_top + AHB-Lite memory models + "
               "IRQ stimulus. Self-checking directed sequences, no software "
               "toolchain required yet.", SANITY)
    note(ws, ["Prerequisite: AHB-Lite behavioural memory models with "
              "programmable wait states and ERROR injection.",
              "  The strategy doc's diagram shows generic I-MEM/D-MEM models; "
              "for GARUDA these must speak AHB-Lite, not OBI or AXI."])

    ws = sheet(wb, "3 Core ISA-ISS", "Phase 3 — Core ISA / ISS Lockstep",
               "Strategy doc §4. riscv-arch-test + riscv-dv against Spike, "
               "compared per retired instruction.", ISS)
    note(ws, ["This phase is where most core bugs will actually be found.",
              "  The retire tag added to mem_wb on 2026-07-20 is the signal a "
              "lockstep harness needs — that work is a prerequisite, not a "
              "counter fix.",
              "Red rows are GARUDA-specific blockers that need a decision "
              "before the phase can start."])

    ws = sheet(wb, "4 SoC no Core", "Phase 4 — SoC Verification without Core",
               "Strategy doc §5. BLOCKED: the SoC RTL does not exist yet.",
               SOC_NO_CORE)
    note(ws, ["The strategy doc notes phase 4 can run in parallel with 1-3.",
              "  For GARUDA it cannot: every SoC RTL directory is empty. "
              "Listed here so the dependency is explicit rather than implied."])

    sheet(wb, "5 Full SoC", "Phase 5 — Full SoC with Core",
          "Strategy doc §6. BLOCKED: depends on Phase 4.", FULL_SOC)

    sheet(wb, "Coverage", "Coverage & Signoff Criteria — Core",
          "The strategy doc defines signoff as coverage closure + bug "
          "burn-down, not 'it ran once'.", CORE_COVERAGE)

    ws = sheet(wb, "Infrastructure", "Infrastructure & Tools — Core",
               "What we have, what we need, and who provides it.", CORE_INFRA)
    note(ws, ["HAVE = available now.  TODO = not installed.  "
              "ASK = needs the coach.  PARTIAL = works with caveats."])

    sheet(wb, "Open Questions", "Open Questions & Decisions — Core",
          "Putting open questions in the submitted plan is deliberate — the "
          "same discipline the project already applies to spec errata.",
          CORE_QUESTIONS)

    path = os.path.abspath(CORE_OUT)
    wb.save(path)
    return wb, path


def build_dsu():
    wb = openpyxl.Workbook()
    sheet(wb, "0 Summary", "GARUDA DSU — Verification Plan",
          "Block II, the collision-avoidance coprocessor. The core has its "
          "own plan: tb/core/GARUDA_Core_Verification_Plan.xlsx",
          DSU_SUMMARY, first=True)

    ws = sheet(wb, "1 DSU Tests", "DSU Test Plan",
               "Supersedes tb/dsu/DSU_Verification_Plan.xlsx: the same rows, "
               "corrected, plus the gaps found by the 2026-07-20 "
               "golden-model audit. Red rows are confirmed RTL findings.", DSU)
    note(ws, ["Reference model: tools/golden/DSU_Golden.py",
              "  DSUModel  = cycle-accurate, bit-literal CSA/adder. Use this "
              "for RTL comparison.",
              "  ArchDSU   = timeless value model. Maths sanity only — it "
              "cannot judge timing.",
              "  Validated against RTL 2026-07-20: FLAG-A, FLAG-B and the "
              "baseline all match.",
              "Legend:",
              "  Pri  P0 = must pass to tape out.  P1 = should pass.",
              "  Done PASS = passes today.  OPEN = confirmed RTL finding.  "
              "FIXED = plan row corrected."])

    sheet(wb, "Coverage", "Coverage & Signoff Criteria — DSU",
          "Signoff is coverage closure plus FLAG closure, not 'the vectors "
          "passed'.", DSU_COVERAGE)

    ws = sheet(wb, "Infrastructure", "Infrastructure & Tools — DSU",
               "The DSU is the only block with no external reference model, "
               "so we own more of this stack than usual.", DSU_INFRA)
    note(ws, ["HAVE = available now.  TODO = not built.  "
              "ASK = needs the coach.  PARTIAL = works with caveats."])

    sheet(wb, "Open Questions", "Open Questions & FLAGs — DSU",
          "FLAG-A and FLAG-C are confirmed RTL bugs, reproduced on real "
          "hardware under simulation. They need a decision before RTL freeze.",
          DSU_QUESTIONS)

    path = os.path.abspath(DSU_OUT)
    wb.save(path)
    return wb, path


def main():
    for wb, path in (build_core(), build_dsu()):
        print(f"wrote {path}")
        for ws in wb:
            print(f"    {ws.title:<18} {ws.max_row - 4} rows")


if __name__ == "__main__":
    main()
